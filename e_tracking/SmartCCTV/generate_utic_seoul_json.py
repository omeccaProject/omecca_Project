"""
generate_utic_seoul_json.py
==============================================
OpenDataCCTV.xlsx(전국 CCTV) -> web/data/utic-cameras-seoul.json(서울만)

필터링 기준
----------------------------------------------------
1차 기준: CENTERNAME == '서울교통정보센터'
  - 단순히 CCTV 이름에 "서울"이 포함되는지로 판단하지 않는다.
  - 실제 데이터 확인 결과, CENTERNAME은 전국 35개 제공기관으로 명확히
    구분되어 있고 '서울교통정보센터'가 정확히 서울시 CCTV를 가리킨다.
    (예: '이수교차로'/천안교통정보센터 는 이름에 "이수"가 들어가지만
     서울이 아니므로 제외되어야 하고, 실제로 CENTERNAME 기준으로 제외됨)

2차 검증: 좌표가 서울시 대략적 경계(위도 37.40~37.75, 경도 126.70~127.25)
  안에 있는지 확인한다. (1차 기준으로 걸러진 303건 전부 이 범위 안에
  있음을 이미 확인했지만, 원본 데이터가 바뀌어도 안전하도록 유지한다)

좌표 처리
----------------------------------------------------
XCOORD = longitude(경도) -> lng
YCOORD = latitude(위도)  -> lat
(반대로 바꾸지 않는다)
"""

import json
import openpyxl

SOURCE_XLSX = "/mnt/user-data/uploads/OpenDataCCTV.xlsx"
OUTPUT_JSON = "/home/claude/web/data/utic-cameras-seoul.json"

SEOUL_CENTER_NAME = "서울교통정보센터"

# 2차 검증용 서울시 대략적 경계 (여유를 둔 값)
LAT_MIN, LAT_MAX = 37.40, 37.75
LNG_MIN, LNG_MAX = 126.70, 127.25


def main():
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb["sheet2"]

    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert header == ["RN", "CCTVID", "CCTVNAME", "CENTERNAME", "XCOORD", "YCOORD"], (
        f"예상한 컬럼 구성과 다릅니다: {header}"
    )

    total_count = 0
    seoul_cameras = []
    excluded_count = 0
    coord_error_count = 0

    for r in range(2, ws.max_row + 1):
        rn = ws.cell(row=r, column=1).value
        cctv_id = ws.cell(row=r, column=2).value
        cctv_name = ws.cell(row=r, column=3).value
        center_name = ws.cell(row=r, column=4).value
        x_coord = ws.cell(row=r, column=5).value  # 경도
        y_coord = ws.cell(row=r, column=6).value  # 위도

        if cctv_id is None:
            continue  # 빈 행 스킵
        total_count += 1

        if center_name != SEOUL_CENTER_NAME:
            excluded_count += 1
            continue

        try:
            lng = float(x_coord)
            lat = float(y_coord)
        except (TypeError, ValueError):
            coord_error_count += 1
            excluded_count += 1
            continue

        if not (LAT_MIN <= lat <= LAT_MAX and LNG_MIN <= lng <= LNG_MAX):
            # CENTERNAME은 서울인데 좌표가 서울 범위를 벗어나는 이상 데이터 -> 안전하게 제외
            coord_error_count += 1
            excluded_count += 1
            continue

        seoul_cameras.append(
            {
                "cam_id": cctv_id,
                "cctv_id": cctv_id,
                "name": cctv_name,
                "center_name": center_name,
                "lng": lng,
                "lat": lat,
                "source": "UTIC",
                "rn": rn,  # 원본 순번 (참고/추적용, 검색에는 사용하지 않음)
            }
        )

    seoul_cameras.sort(key=lambda c: (c["name"] or ""))

    payload = {
        "description": "서울교통정보센터(UTIC) 제공 서울 지역 실시간 CCTV 좌표 데이터",
        "source_file": "OpenDataCCTV.xlsx",
        "filter_criteria": (
            f"CENTERNAME == '{SEOUL_CENTER_NAME}' "
            f"AND {LAT_MIN} <= lat <= {LAT_MAX} AND {LNG_MIN} <= lng <= {LNG_MAX}"
        ),
        "total_source_count": total_count,
        "seoul_count": len(seoul_cameras),
        "excluded_count": excluded_count,
        "coord_error_count": coord_error_count,
        "cameras": seoul_cameras,
    }

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"전체 원본 레코드: {total_count}")
    print(f"서울(UTIC) 레코드: {len(seoul_cameras)}")
    print(f"제외된 레코드: {excluded_count} (좌표 이상: {coord_error_count})")
    print(f"저장 위치: {OUTPUT_JSON}")

    # 검증: L010263 이수역이 정확히 들어갔는지 확인
    target = next((c for c in seoul_cameras if c["cam_id"] == "L010263"), None)
    print()
    if target:
        print("검증(L010263 이수역):", json.dumps(target, ensure_ascii=False))
        assert target["name"] == "이수역"
        assert target["center_name"] == "서울교통정보센터"
        assert abs(target["lat"] - 37.4846) < 1e-6
        assert abs(target["lng"] - 126.9824) < 1e-6
        print("-> L010263 검증 통과 (이름/센터/좌표 모두 정확히 일치)")
    else:
        print("!! 경고: L010263을 서울 데이터에서 찾지 못했습니다.")


if __name__ == "__main__":
    main()